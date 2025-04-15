import os
import sys
import traceback
from openpyxl import load_workbook

import os
import traceback
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def get_excel(route):
    """
    Opens an Excel workbook and returns the workbook object and the specified worksheet.

    Parameters:
    sheet_name (str): Name of the worksheet to access
    route (str): File path to the Excel workbook

    Returns:
    tuple: (workbook_object, worksheet_object) if successful
           (None, None) if any error occurs
    """
    try:
        # Verify file exists
        if not os.path.exists(route):
            print(f"ERROR: The Excel file does not exist at path: {route}")
            return None, None

        # Verify file extension
        if not route.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
            print(f"ERROR: The file {route} does not have a valid Excel extension (.xlsx, .xlsm, .xltx, .xltm)")
            return None, None

        # Check file size (if too small, likely not a valid Excel file)
        file_size = os.path.getsize(route)
        if file_size < 2000:  # Excel files are typically at least a few KB
            print(
                f"WARNING: The file {route} is suspiciously small ({file_size} bytes). It may not be a valid Excel file.")

        print(f"Opening Excel file: {route}")
        wb = load_workbook(filename=route, read_only=False, keep_vba=False)




        return wb

    except InvalidFileException:
        print(f"ERROR: The file {route} is not a valid Excel file or is corrupted.")
        print("Try opening and resaving the file in Excel to fix potential corruption issues.")
        return None
    except KeyError as e:
        print(f"ERROR: Excel file appears to be corrupt or in an incompatible format: {str(e)}")
        print("The file might be in a format not supported by openpyxl (like .xls) or may be corrupt.")
        return None
    except PermissionError:
        print(
            f"ERROR: Cannot access {route}. The file may be open in another application or you lack permission to access it.")
        return None, None
    except Exception as e:
        print(f"ERROR in get_excel: {str(e)}")
        traceback.print_exc()
        return None

