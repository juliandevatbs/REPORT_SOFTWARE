import os
import sys
import traceback
from openpyxl import load_workbook

def get_sheet_names(route: str) -> list:
    """
    Retrieves all sheet names from an Excel workbook using openpyxl.
    
    Args:
        route (str): Path to the Excel workbook file
        
    Returns:
        list: A list of sheet names if successful, or None if an error occurs.
    """
    wb = None
    
    try:
        # Check if the file exists
        if not os.path.exists(route):
            raise FileNotFoundError(f"The file {route} does not exist")
        
        # Open the workbook in read-only mode (faster and uses less memory)
        wb = load_workbook(filename=route, read_only=True)
        
        # Get all sheet names
        sheet_names = wb.sheetnames
        
        return sheet_names

    except Exception as e:
        print(f"Error: {str(e)}", file=sys.stderr)
        traceback.print_exc()
        return None
        
    finally:
        # Ensure workbook is closed properly
        if wb is not None:
            try:
                wb.close()
            except Exception as e:
                print(f"Error closing workbook: {str(e)}", file=sys.stderr)