import re
from openpyxl.cell import MergedCell
from openpyxl.utils import column_index_from_string


def write_cell(ws, celda_coord, valor) -> bool:
    """
    Write a value to a specified cell in an Excel worksheet, handling merged cells.

    This function writes a value to a specified cell coordinate in an Excel worksheet.
    When the target cell is part of a merged range, the function identifies the primary cell
    (top-left cell) of the merged range and writes the value there, maintaining proper Excel behavior.

    Args:
        ws: The worksheet object (openpyxl Worksheet)
        celda_coord (str): Cell coordinate in Excel format (e.g., 'A1', 'B12')
        valor: The value to write to the cell

    Returns:
        bool: True if the write operation was successful, False otherwise

  """
    try:
        # Parse the cell coordinate using regex to extract column letters and row number
        match = re.match(r'([A-Za-z]+)(\d+)', celda_coord)
        if not match:
            print(f"Formato de coordenada incorrecto: {celda_coord}")
            return False

        # Extract column letters and row number from the match
        col_str, row_str = match.groups()
        row = int(row_str)
        col = column_index_from_string(col_str)

        # Get the cell at the specified coordinate
        celda = ws.cell(row=row, column=col)

        # If it's a regular cell (not merged), write the value directly
        if not isinstance(celda, MergedCell):
            celda.value = valor
            return True

        # If it's a merged cell, find the primary cell in the merged range
        for rango in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = rango.min_row, rango.min_col, rango.max_row, rango.max_col

            # Check if the target cell is within this merged range
            if min_row <= row <= max_row and min_col <= col <= max_col:
                # Get the primary cell (top-left) of the merged range
                celda_principal = ws.cell(row=min_row, column=min_col)
                # Write the value to the primary cell
                celda_principal.value = valor
                return True

        # If no matching merged range was found (unusual case)
        print(f"No se encontró un rango combinado para la celda {celda_coord}")
        return False

    except Exception as e:
        # Catch any unexpected errors that might occur
        print(f"Error en write_cell con coordenada {celda_coord}: {str(e)}")
        return False