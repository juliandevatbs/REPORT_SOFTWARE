def clean_cell(ws, cell_coord):
    """
    Limpia el contenido de una celda, incluso si es parte de un rango combinado.

    Args:
        ws: Hoja de trabajo de openpyxl
        cell_coord: Coordenada de la celda (ej: 'B13')

    Returns:
        bool: True si se pudo limpiar, False en caso contrario
    """
    try:
        from openpyxl.utils import column_index_from_string
        import re

        match = re.match(r'([A-Za-z]+)(\d+)', cell_coord)
        if not match:
            return False

        col_str, row_str = match.groups()
        row = int(row_str)
        col = column_index_from_string(col_str)

        # Obtener la celda
        cell = ws.cell(row=row, column=col)

        # Verificar si la celda está en un rango combinado (nueva forma para openpyxl 3.0+)
        is_merged = False
        main_cell = None

        for merged_range in ws.merged_cells:
            if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= col <= merged_range.max_col):
                is_merged = True
                main_cell = ws.cell(row=merged_range.min_row,
                                    column=merged_range.min_col)
                break

        if is_merged and main_cell:
            # Limpiar la celda principal del rango combinado
            main_cell.value = ''
        else:
            # Limpiar celda normal
            cell.value = ''

        return True

    except Exception as e:
        print(f"Error cleaning cell {cell_coord}: {str(e)}")
        return False