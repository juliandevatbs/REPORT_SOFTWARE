import os
import subprocess
import sys
import time
import traceback

from openpyxl.reader.excel import load_workbook


def safe_save_workbook(wb, route, max_attempts=3):

    """Intenta guardar el workbook con reintentos y manejo de errores"""
    for attempt in range(max_attempts):

        try:
            # Cerrar Excel antes de guardar
            kill_excel_processes()
            time.sleep(1)  # Esperar para asegurar cierre

            # Guardar con backup
            temp_route = route + ".temp"
            wb.save(temp_route)

            # Reemplazar archivo original
            if os.path.exists(route):
                os.remove(route)
                os.rename(temp_route, route)

            print(f"Workbook saved successfully on attempt {attempt + 1}")
            return True
        except Exception as e:
            print(f"Intento {attempt + 1} de guardar falló: {str(e)}")
            time.sleep(2)  # Wait before retrying
        return False

def kill_excel_processes():
    """Cierra todos los procesos de Excel"""
    try:
        subprocess.run(["taskkill", "/f", "/im", "excel.exe"],
                       stdout=subprocess.DEVNULL,
                       stderr=subprocess.DEVNULL)
        time.sleep(1)

    except:
        pass

def clean_constant_data():
    """Limpia los datos analíticos en la hoja de reporte"""
    try:
        # Configuración
        sheetname = "Reporte"
        route = r"C:/Users/Duban Serrano/Desktop/REPORTES PYTHON/excel/Reporte 2025-03-12 (3).xlsx"
        start_row = 114
        row_spacing = 5
        num_blocks = 13

        # Verificar si el archivo existe
        if not os.path.exists(route):
            print(f"Error: Archivo no encontrado en {route}")
            return False

        # Abrir el archivo Excel
        print("Abriendo workbook...")
        try:
            wb = load_workbook(filename=route)
            ws = wb[sheetname]
        except Exception as e:
            print(f"Error al abrir el archivo: {str(e)}")
            return False

        # Limpiar datos
        print(f"Limpiando hasta {num_blocks} bloques de datos...")
        for block_num in range(num_blocks):

            first_line_row = start_row + (block_num * row_spacing)
            second_line_row = first_line_row + 2

            try:
                # Celdas de la primera línea a limpiar
                first_line_cells = [
                    f"J{first_line_row}",  # sw_code
                    f"J{first_line_row}",  # batch_id_value
                    f"R{first_line_row}",  # date_value_str
                    f"Z{first_line_row}",  # by_value
                    f"AJ{first_line_row}"  # matrix_id_value
                ]

                # Celdas de la segunda línea a limpiar
                second_line_cells = [
                    f"J{second_line_row}",  # result_value
                    f"AD{second_line_row}",  # date_value_str
                    f"AF{second_line_row}",  # by_value
                    f"AH{second_line_row}"  # batch_id_value
                ]

                third_line_cells = [

                ]

                # Limpiar todas las celdas
                for cell_ref in first_line_cells + second_line_cells:
                    print(cell_ref)
                    ws[cell_ref].value = ''
                    # Opcional: resetear el formato si es necesario
                    # ws[cell_ref]._style = None

                print(f"Bloque {block_num} limpiado en filas {first_line_row}-{second_line_row}")

            except Exception as block_error:
                print(f"Error al limpiar bloque {block_num}: {str(block_error)}")
                continue

        # Guardar cambios
        print("Guardando workbook...")
        if not safe_save_workbook(wb, route):
            print("Error: No se pudo guardar el archivo después de varios intentos")
            return False

        print("Archivo guardado exitosamente con las celdas limpiadas")
        return True

    except Exception as e:
        print(f"Error crítico: {str(e)}")
        traceback.print_exc()
        return False
    finally:
        if 'wb' in locals():
            wb.close()
        kill_excel_processes()