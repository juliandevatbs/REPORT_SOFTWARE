import os
import subprocess
import sys
import time

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def safe_save_workbook(wb, route, max_attempts=3):
    """Intenta guardar el workbook con reintentos y manejo de errores"""
    for attempt in range(max_attempts):

        try:
            # Cerrar Excel antes de guardar
            kill_excel_processes()
            time.sleep(1)  # Esperar para asegurar cierre

            # Guardar con backup
            temp_route = route + ".temp"
            wb.save(temp_route)

            # Reemplazar archivo original
            if os.path.exists(route):
                os.remove(route)
                os.rename(temp_route, route)

            print(f"Workbook saved successfully on attempt {attempt + 1}")
            return True
        except Exception as e:
            print(f"Intento {attempt + 1} de guardar falló: {str(e)}")
            time.sleep(2)  # Wait before retrying
        return False


def kill_excel_processes():
    """Cierra todos los procesos de Excel"""
    try:
        subprocess.run(["taskkill", "/f", "/im", "excel.exe"],
                       stdout=subprocess.DEVNULL,
                       stderr=subprocess.DEVNULL)
        time.sleep(1)
    except:
        pass


def clean_lab_data():
        sheet_name = "Reporte"
        route = r"C:\Users\Duban Serrano\Desktop\REPORTES PYTHON\excel\Reporte 2025-03-12 (3).xlsx"

        wb = None

        try:
            # Verify file exists and is accessible
            if not os.path.exists(route):
                print(f"Error: File not found at {route}")
                return False

            # Open workbook with different options
            wb = load_workbook(filename=route, read_only=False, data_only=True)

            # Verify sheet exists
            if sheet_name not in wb.sheetnames:
                print(f"Error: Worksheet '{sheet_name}' not found")
                return False

            ws = wb[sheet_name]

            # Cells to clear with comprehensive clearing strategy
            columns_to_clear = [
                'B', 'G', 'K', 'Q', 'U', 'X', 'AC'
            ]

            rows_to_clear =  20
            start_row = 13

            # Advanced clearing strategy
            for row in range(rows_to_clear):



                for column in columns_to_clear:

                    cell = ws[f"{column}{start_row}"]
                    cell.value = ''
                start_row += 1



            # Disable Excel protection if present
            if hasattr(ws, 'protection'):
                ws.protection.sheet = False

            # Save the workbook using safe save method
            save_success = safe_save_workbook(wb, route)
            if not save_success:
                print("Failed to save the workbook after multiple attempts")
                return False

            print("Cells cleared and workbook saved successfully")
            return True

        except Exception as e:
            print(f"Unexpected error occurred: {e}")
            import traceback
            # This will print the full traceback
            traceback.print_exc()
            return False

        finally:
            # Clean up resources
            if wb is not None:
                try:
                    wb.close()
                except Exception as e:
                    print(f"Error closing workbook: {e}")

    # Run the function
if __name__ == "__main__":
    result = clean_lab_data()
    print(f"Script execution result: {'Success' if result else 'Failure'}")
    sys.exit(0 if result else 1)