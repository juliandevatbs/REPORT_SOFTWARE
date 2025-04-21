import traceback
import subprocess
from openpyxl import load_workbook
from datetime import datetime
import sys
import os
import time
import re
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter

# Configuración de rutas
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from scripts.get.get_all_data import *
from scripts.utils.safe_save import safe_save_workbook


def group_by_sheet(row_data):

    """Agrupa los datos por nombre de hoja (primer elemento de cada sublista)"""
    sheets_dict = {}

    for row in row_data:
        for ro in row:
            print(ro)

def write_cell(ws, celda_coord, valor):
    """
    Escribe un valor en una celda, incluso si es parte de un rango combinado.

    Args:
        ws: Hoja de trabajo
        celda_coord: Coordenada de la celda (ej: 'B13')
        valor: Valor a escribir

    Returns:
        bool: True si se pudo escribir, False en caso contrario
    """
    try:

        match = re.match(r'([A-Za-z]+)(\d+)', celda_coord)
        if not match:
            return False

        col_str, row_str = match.groups()
        row = int(row_str)
        col = column_index_from_string(col_str)

        # Intentar obtener la celda directamente
        celda = ws.cell(row=row, column=col)

        # Si no es una celda combinada, escribir directamente
        if not isinstance(celda, MergedCell):
            celda.value = valor
            return True

        for rango in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = rango.min_row, rango.min_col, rango.max_row, rango.max_col

            if min_row <= row <= max_row and min_col <= col <= max_col:
                celda_principal = ws.cell(row=min_row, column=min_col)
                celda_principal.value = valor
                return True

        return False

    except Exception as e:
        return False





def should_write_block(data_block):
    """Determina si el bloque debe escrib12irse según la condición data_block[18] > data_block[12]"""
    try:

        value_result = float(data_block[18]) if data_block[18] is not None else None
        value_pql = float(data_block[14]) if data_block[14] is not None else None
        
        print(f"RESULT{value_result}-----------PQLLLL{value_pql}")
        if value_result is None or value_pql is None:
            return False

        return value_result > value_pql
    except (ValueError, TypeError):
        return False

def validate_data_block(data_block):
    if not isinstance(data_block, list):
        return False
    return True

def format_date(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return str(value) if value is not None else ""


def write_data_block(ws, data_block, first_line_row):
    try:
        if not validate_data_block(data_block):
            return False

        sw_code = str(data_block[1]) if data_block[1] is not None else ""
        date_value = f"{format_date(data_block[2])} {data_block[3]}"
        by_value = str(data_block[8]) if data_block[8] is not None else ""
        result_value = data_block[18] if data_block[18] is not None else ""
        batch_id_value = data_block[7] if data_block[7] is not None else ""
        matrix_id_value = data_block[5] if data_block[5] is not None else ""
        df_value = data_block[13]
        mdl_value = data_block[14]
        pql_value = data_block[15]
        units_value = data_block[16]
        analyzed_method = data_block[10]
        analyte_name = data_block[9]
        notes = data_block[19]

        second_line_row = first_line_row + 2

        cell_mapping = {
            f"B{first_line_row}": sw_code,
            f"J{first_line_row}": batch_id_value,
            f"R{first_line_row}": date_value,
            f"Z{first_line_row}": by_value,
            f"AJ{first_line_row}": matrix_id_value,
            f"B{second_line_row}": analyte_name,
            f"J{second_line_row}": result_value,
            f"AD{second_line_row}": date_value,
            f"AF{second_line_row}": by_value,
            f"AH{second_line_row}": batch_id_value,
            f"R{second_line_row}": units_value,
            f"U{second_line_row}": df_value,
            f"V{second_line_row}": mdl_value,
            f"W{second_line_row}": pql_value,
            f"Z{second_line_row}": analyzed_method,
            f"AJ{second_line_row}": ''
        }

        for cell, value in cell_mapping.items():
            write_cell(ws, cell, value)

        return True

    except Exception as e:
        print(f"ERROR FATAL EN {data_block[9]}: {str(e)}")
        traceback.print_exc()
        return False


def print_summary_data(ws, row_data, current_row: int):
    try:

        if not row_data:
            print("No hay datos para escribir")
            return False

        grouped_data = {}
        ordered_analytes = []

        for data_block in row_data:
            print(" DATAAAAAAA BLOCKSSSS")
            print(data_block)
            if validate_data_block(data_block) and should_write_block(data_block):
                analyte_name = data_block[9] or "Sin Nombre"
                if analyte_name not in grouped_data:
                    grouped_data[analyte_name] = []
                    ordered_analytes.append(analyte_name)
                grouped_data[analyte_name].append(data_block)

        # El resto del código es igual que la función original
        success_count = 0
        row_spacing = 4

        for analyte in ordered_analytes:
            for data_block in grouped_data[analyte]:
                print(len(data_block))
                print(data_block[9])
                print(data_block[18])
                if write_data_block(ws, data_block, current_row):
                    success_count += 1
                    current_row += row_spacing
                    print(f"Escrito bloque para analito: {analyte} en fila {current_row}")
                else:
                    print(f"Error escribiendo bloque para analito: {analyte}")

        print(f"Bloques escritos exitosamente: {success_count}/{len(row_data)}")
        return True

    except Exception as e:
        print(f"Error crítico: {str(e)}")
        traceback.print_exc()
        return False

