import re
import sys
import os
import time
import psutil
import traceback
from openpyxl.cell import MergedCell
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string

from scripts.blocks.test import copy_range_with_styles

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from scripts.excel.connect_excel import get_excel
from scripts.get.get_header_data import get_header_data
from scripts.error.show_error import show_info
from scripts.utils.kill_excel_processes import kill_excel_processes

def check_file_locks(filepath):
    """Verifica si el archivo está bloqueado por otro proceso"""
    for proc in psutil.process_iter():
        try:
            files = proc.open_files()
            for f in files:
                if filepath.lower() == f.path.lower():
                    return True
        except (psutil.AccessDenied, psutil.NoSuchProcess):
            continue
    return False


def write_cell(ws, celda_coord, valor):
    """Versión mejorada con mejor manejo de errores"""
    try:
        match = re.match(r'([A-Za-z]+)(\d+)', celda_coord)
        if not match:
            return False

        col_str, row_str = match.groups()
        row = int(row_str)
        col = column_index_from_string(col_str)

        celda = ws.cell(row=row, column=col)

        if not isinstance(celda, MergedCell):
            celda.value = valor
            return True

        for rango in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = rango.min_row, rango.min_col, rango.max_row, rango.max_col
            if min_row <= row <= max_row and min_col <= col <= max_col:
                ws.cell(row=min_row, column=min_col).value = valor
                return True

        return False
    except Exception as e:
        print(f"Error writing cell {celda_coord}: {str(e)}")
        return False


def print_header_data(wb, ws, header_data, spacing_data: list):
    """Versión mejorada con manejo robusto de archivos"""
    try:

        # Procesar datos
        na_value = 'No Application'
        fields = {
            "company_name": str(header_data[0] if header_data[0] else na_value),
            "client_name": str(header_data[1] if header_data[1] else na_value),
            "client_address": str(header_data[2] if header_data[2] else na_value),
            "city": str(header_data[3] if header_data[3] else na_value),
            "state": str(header_data[4] if header_data[4] else na_value),
            "zip_code": str(header_data[5] if header_data[5] else na_value),
            "facility_id": str(header_data[6] if header_data[6] else na_value),
            "requested_data": str(header_data[7] if header_data[7] else na_value),
            "project_location": str(header_data[8] if header_data[8] else na_value),
            "client_phone": str(header_data[9] if header_data[9] else na_value),
            "project_number": str(header_data[10] if header_data[10] else na_value),
            "lab_reporting_batch_id": str(header_data[11] if header_data[11] else na_value)

        }

        cell_mapping = {

            "company_name": [f"G{spacing_data[0]}", f"G{spacing_data[0]}", f"G{spacing_data[0]}",
                             f"G{spacing_data[0]}"],
            "client_name": [f"G{spacing_data[1]}", f"G{spacing_data[1]}", f"G{spacing_data[1]}", f"G{spacing_data[1]}",
                            f"G{spacing_data[1]}"],
            "client_address": [f"G{spacing_data[2]}", f"G{spacing_data[2]}", f"G{spacing_data[2]}", f"G{spacing_data[2]}",
                               f"G{spacing_data[2]}"],
            "city": [f"G{spacing_data[3]}", f"G{spacing_data[3]}", f"G{spacing_data[3]}", f"G{spacing_data[3]}",
                     f"G{spacing_data[3]}"],
            "state": [f"G{spacing_data[4]}", f"G{spacing_data[4]}", f"G{spacing_data[4]}", f"G{spacing_data[4]}",
                      f"G{spacing_data[4]}"],
            "zip_code": [f"L{spacing_data[5]}", f"L{spacing_data[5]}", f"L{spacing_data[5]}", f"L{spacing_data[5]}", f"L{spacing_data[5]}"],
            "requested_data": [f"AF{spacing_data[6]}", f"AF{spacing_data[6]}",
                               f"AF{spacing_data[6]}", f"AF{spacing_data[6]}",
                               f"AF{spacing_data[6]}"],
            "facility_id": [f"AF{spacing_data[7]}", f"AF{spacing_data[7]}", f"AF{spacing_data[7]}",
                            f"AF{spacing_data[7]}", f"AF{spacing_data[7]}"],
            "project_location": [f"AF{spacing_data[8]}", f"AF{spacing_data[8]}",
                                 f"AF{spacing_data[8]}", f"AF{spacing_data[8]}",
                                 f"AF{spacing_data[8]}"],
            "client_phone": [f"AF{spacing_data[9]}", f"AF{spacing_data[9]}", f"AF{spacing_data[9]}",
                             f"AF{spacing_data[9]}", f"AF{spacing_data[9]}"],
            "project_number": [f"AF{spacing_data[10]}", f"AF{spacing_data[10]}",
                               f"AF{spacing_data[10]}", f"AF{spacing_data[10]}",
                               f"AF{spacing_data[10]}"],
            "lab_reporting_batch_id": [f"AF{spacing_data[11]}", f"AF{spacing_data[11]}",
                                       f"AF{spacing_data[11]}", f"AF{spacing_data[11]}",
                                       f"AF{spacing_data[11]}"]

        }


        for field, cells in cell_mapping.items():
            for cell in cells:
                if not write_cell(ws, cell, fields[field]):
                    print(f"Error escribiendo {field} en {cell}")

    except Exception as e:
        show_info(f"Error crítico: {str(e)}")
        traceback.print_exc()
        return False

