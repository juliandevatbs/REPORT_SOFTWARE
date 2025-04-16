import traceback
import subprocess
from openpyxl import load_workbook
from datetime import datetime
import sys
import os
import time
import re
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string

from scripts.get.get_all_data import get_chain_data
from scripts.get.get_quality_data import get_q_data

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))


def write_cell(ws, celda_coord, valor):
    """
    Escribe un valor en una celda, incluso si es parte de un rango combinado.
    """
    try:
        match = re.match(r'([A-Za-z]+)(\d+)', celda_coord)
        if not match:
            print(f"Incorrect format: {celda_coord}")
            return False

        col_str, row_str = match.groups()
        row = int(row_str)
        col = column_index_from_string(col_str)
        celda = ws.cell(row=row, column=col)

        if not isinstance(celda, MergedCell):
            celda.value = valor
            return True

        for rango in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = rango.min_row, rango.min_col, rango.max_row, rango.max_col
            if min_row <= row <= max_row and min_col <= col <= max_col:
                ws.cell(row=min_row, column=min_col).value = valor
                return True

        return False

    except Exception as e:
        print(f"Write in the cell failed: {str(e)}")
        return False


def print_quality_data(route: str, last_lab_sample_id):
    try:
        sheetname = "Reporte"
        start_row = 231
        row_spacing = 7

        # Obtener datos con ID secuencial
        row_data = get_q_data(base_sample_id=last_lab_sample_id)

        if not row_data:
            print("Error: There is no data to write")
            return False

        if not os.path.exists(route):
            print(f"File not found: {route}")
            return False

        print("Open WorkBook...")
        try:
            wb = load_workbook(filename=route)
            ws = wb[sheetname]
        except Exception as e:
            print(f"Error opening file: {str(e)}")
            return False

        for block_num, data_block in enumerate(row_data):
            print(f"Processing block {block_num}: {data_block}")
            if not isinstance(data_block, list) or len(data_block) < 16:
                print(f"Invalid data block: {data_block}")
                continue

            try:
                # Estructura actual de data_block:
                # [sample_id, sheet_name, B, C, D, E, F, G, H, I, J, K, N19, N20, N21, N22, N23, N24]

                first_line_row = start_row + (block_num * row_spacing)
                second_line_row = first_line_row + 1
                third_line_row = first_line_row + 2
                fourth_line_row = first_line_row + 3
                fifth_line_row = first_line_row +5


                print(first_line_row, second_line_row, third_line_row, fourth_line_row, fifth_line_row)

                # Asignación corregida según nueva estructura
                client_sample_id = data_block[0]  # sample_id generado
                sheet_name = data_block[1]  # nombre de hoja
                sampled = data_block[3]  # valor columna B (fecha)
                lab_sample = data_block[3]  # valor columna C (tipo de muestra)
                analyte_name = sheet_name.split('(')[0].strip()  # extraer nombre de analito
                results = data_block[8]  # valor columna I
                units = data_block[17]  # último valor constante (N24)
                df = data_block[13]  # N20
                mdl = data_block[14]  # N21
                pql = data_block[15]  # N22
                prep = sampled  # misma fecha que sampled
                analyzed = "EPA 9212"
                matrix_id = "GroundWater"
                by = ""
                batch_id = f"{client_sample_id}{lab_sample}"
                date =data_block[2]

                # Primera línea
                write_cell(ws, f"B{first_line_row}", sampled)

                # Segunda línea
                write_cell(ws, f"J{second_line_row}", client_sample_id)
                write_cell(ws, f"AC{second_line_row}", date)

                # Tercera linea
                write_cell(ws, f"J{third_line_row}", lab_sample)
                write_cell(ws, f"AC{third_line_row}", prep)
                write_cell(ws, f"AE{third_line_row}", analyzed)
                write_cell(ws, f"AJ{third_line_row}", matrix_id)

                write_cell(ws, f"B{fourth_line_row}", sampled)

                # Cuarta linea
                write_cell(ws, f"B{fifth_line_row}", analyte_name)
                write_cell(ws, f"J{fifth_line_row}", results)
                write_cell(ws, f"O{fifth_line_row}", units)
                write_cell(ws, f"T{fifth_line_row}", df)
                write_cell(ws, f"U{fifth_line_row}", mdl)
                write_cell(ws, f"V{fifth_line_row}", pql)
                write_cell(ws, f"W{fifth_line_row}", analyzed)
                write_cell(ws, f"AD{fifth_line_row}", date)
                write_cell(ws, f"AH{fifth_line_row}", by)
                write_cell(ws, f"AJ{fifth_line_row}", analyzed)

            except Exception as block_error:
                print(f"Error en bloque {block_num}: {str(block_error)}")
                print(f"Datos problemáticos: {data_block}")
                traceback.print_exc()
                continue

        # Guardar los cambios
        try:
            #wb.save(route)
            print("Datos escritos exitosamente")
            return True
        except Exception as e:
            print(f"Error al guardar el archivo: {str(e)}")
            return False

    except Exception as e:
        print(f"Error crítico: {str(e)}")
        traceback.print_exc()
        return False

