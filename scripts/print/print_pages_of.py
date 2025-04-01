from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import os


def agregar_numero_paginas(excel_path, celda="AK35", hoja="Reporte"):
    """
    Agrega "Página 1 de X" en una celda específica.
    :param excel_path: Ruta del archivo Excel.
    :param celda: Celda donde se insertará el texto (ej: "A1").
    :param hoja: Nombre de la hoja a modificar.
    """
    try:
        # Cargar el workbook
        wb = load_workbook(excel_path)
        ws = wb[hoja]

        # Obtener el número total de páginas (estimado)
        # OpenPyXL no puede calcularlo exactamente, usamos un valor fijo o lógica propia
        total_paginas = 3  # Reemplaza con tu valor real o lógica para calcularlo

        # Insertar texto en la celda
        ws[celda] = f"Página 1 de {total_paginas}"

        # Formato (opcional)
        ws[celda].font = Font(bold=True, size=12)
        ws[celda].alignment = Alignment(horizontal="center")

        # Guardar cambios
        wb.save(excel_path)
        print(f"✓ Texto agregado en {celda}: {ws[celda].value}")
        return True

    except Exception as e:
        print(f"✗ Error: {str(e)}")
        return False


# Uso
agregar_numero_paginas(
    excel_path=r"C:\Users\TuUsuario\Desktop\Reporte.xlsx",
    celda="A1",  # Cambia a la celda que necesites
    hoja="Reporte"
)