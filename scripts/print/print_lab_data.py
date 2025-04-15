from datetime import datetime, timedelta
import sys
import os
import traceback
import time
import subprocess
import re
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter

from scripts.blocks.test import copy_range_with_styles
from scripts.excel.connect_excel import get_excel
from scripts.utils.format_date import format_date
from scripts.utils.kill_excel_processes import kill_excel_processes

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from scripts.get.get_all_data import get_chain_data


def write_cell(ws, celda_coord, valor):
    """
    Escribe un valor en una celda, incluso si es parte de un rango combinado.

    Args:
        ws: Hoja de trabajo
        celda_coord: Coordenada de la celda (ej: 'B13')
        valor: Valor a escribir

    Returns:
        bool: True si se pudo escribir, False en caso contrario
    """
    try:

        match = re.match(r'([A-Za-z]+)(\d+)', celda_coord)
        if not match:
            return False

        col_str, row_str = match.groups()
        row = int(row_str)
        col = column_index_from_string(col_str)

        # Intentar obtener la celda directamente
        celda = ws.cell(row=row, column=col)

        # Si no es una celda combinada, escribir directamente
        if not isinstance(celda, MergedCell):
            celda.value = valor
            return True

        for rango in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = rango.min_row, rango.min_col, rango.max_row, rango.max_col

            if min_row <= row <= max_row and min_col <= col <= max_col:
                celda_principal = ws.cell(row=min_row, column=min_col)
                celda_principal.value = valor
                return True

        return False

    except Exception as e:
        return False
        celda = ws.cell(row=row, column=col)

        if not isinstance(celda, MergedCell):
            celda.value = valor
            return True

        for rango in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = rango.min_row, rango.min_col, rango.max_row, rango.max_col
            if min_row <= row <= max_row and min_col <= col <= max_col:
                celda_principal = ws.cell(row=min_row, column=min_col)
                celda_principal.value = valor
                if isinstance(valor, str) and ":" in valor:
                    celda_principal.number_format = 'HH:MM'
                return True

        print(f"No se encontró un rango combinado para la celda {celda_coord}")
        return False

    except Exception as e:
        print(f"Error en write_cell con coordenada {celda_coord}: {str(e)}")
        return False


def print_lab_data(wb, ws, chain_data):
    try:
        print(f"Datos recibidos: {len(chain_data)} filas")
        if not isinstance(chain_data, list) or not all(isinstance(row, list) for row in chain_data):
            print("Error: Datos en formato incorrecto.")
            return False

        start_row = 15
        for row_data in chain_data:

            write_cell(ws, f"B{start_row}", row_data[0])
            write_cell(ws, f"H{start_row}", row_data[7])
            write_cell(ws, f"K{start_row}", row_data[1])
            write_cell(ws, f"U{start_row}", row_data[2])
            write_cell(ws, f"W{start_row}", row_data[3])
            write_cell(ws, f"AC{start_row}", row_data[5])
            write_cell(ws, f"AF{start_row}", row_data[9])
            start_row += 1
        wb.save(r"C:\Users\Duban Serrano\Desktop\REPORTES PYTHON\excel\Reporte 2025-03-12 (4).xlsx")
        print("¡Datos escritos exitosamente!")
        return True

    except Exception as e:
        print(f"Error en print_lab_data: {e}")
        traceback.print_exc()
        return False



