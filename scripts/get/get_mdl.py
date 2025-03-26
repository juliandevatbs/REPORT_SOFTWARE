import os
import sys
from openpyxl import load_workbook

"""

    This function  get the MDL value from the excel 
    
"""

def get_mdl_value():
    
    sheet_name= "Chloride (16887006)"
    route = r"C:/Users/Duban Serrano/Desktop/REPORTES PYTHON/excel/Reporte 2025-03-12 (3).xlsx"
    mdl_value = None
    wb = None
    
    try:
        # Verify file exists
        if not os.path.exists(route):
            print("Error: File not found")
            return None
        
        # Open workbook in read-only mode with data only (no formulas)
        wb = load_workbook(filename=route, read_only=True, data_only=True)
        
        # Verify sheet exists
        if sheet_name not in wb.sheetnames:
            print(f"Error: Worksheet '{sheet_name}' not found")
            return None
        
         # Get worksheet and cell value
        ws = wb[sheet_name]
        mdl_value = ws['N22'].value
        
    except Exception as e:
        print(f"An error occurred: {e}")
        analysis_requested = False
        
    finally:
        # Clean up resources
        if wb is not None:
            try:
                wb.close()
            except Exception as e:
                print(f"Error closing workbook: {e}")
        
        return mdl_value

get_mdl_value()