import os

from openpyxl import load_workbook


def extract_chain_data(route: str):
    """
    Extracts chain of custody data from the 'Chain of custody' sheet,
    only including rows where the Chloride column contains 'yes'.

    Returns:
        list: List of lists containing the filtered row data
    """
    sheetname = 'Chain of Custody 1'
    chloride_column = "B"  # Assuming Chloride data is in column B
    data_columns = ["C", "D", "E", "F"]  # Columns to extract data from
    start_row = 2  # Assuming data starts from row 2 (header in row 1)

    filtered_data = []
    wb = None

    try:
        # Verify file exists
        if not os.path.exists(route):
            print("Error: File not found")
            return []

        # Open workbook in read-only mode
        wb = load_workbook(filename=route, read_only=True, data_only=True)

        # Verify sheet exists
        if sheetname not in wb.sheetnames:
            print(f"Error: Worksheet '{sheetname}' not found")
            return []

        ws = wb[sheetname]

        for row in range(start_row, ws.max_row + 1):
            # Check Chloride column value
            chloride_value = ws[f'{chloride_column}{row}'].value

            # Only process rows where Chloride is 'yes'
            if chloride_value and str(chloride_value).strip().lower() == 'yes':
                # Extract data from specified columns
                row_data = []
                for col in data_columns:
                    row_data.append(ws[f'{col}{row}'].value)
                filtered_data.append(row_data)

    except Exception as e:
        print(f"An error occurred: {e}")
        return []

    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception as e:
                print(f"Error closing workbook: {e}")

    return filtered_data

