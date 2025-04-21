import os
import sys
from openpyxl import load_workbook
import gc

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))


def get_sheet_names(file_path):
    """
    Get all sheet names from an Excel workbook.

    Args:
        file_path (str): Path to the Excel file

    Returns:
        list: List of sheet names
    """
    try:
        wb = load_workbook(filename=file_path, read_only=True, keep_vba=False)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception as e:
        print(f"Error getting sheet names: {e}")
        return []


def get_chain_data(ws):
    """
    Extract chain of custody data from the Excel file.

    Returns:
        list: List of chain data entries
    """
    all_data = []
   

    

    try:
        

        # Verify sheet exists
     

        # Pre-load the constant values
        analysis_requested = ws['AI5'].value
        sampled_by = ws["B10"].value

        # Pre-load the matrix codes dictionary
        matrix_codes = {
            'A': 'Air',
            'GW': 'Groundwater',
            'SE': 'Sediment',
            'SO': 'Soil',
            'SW': 'Surface Water',
            'W': 'Water (Blanks)',
            'HW': 'Potencial Haz Wastw',
            'O': 'Other'
        }

        # Pre-load sheet headers
        sheet_headers = {}
        for col in ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']:
            sheet_headers[col] = {
                'name': ws[f'{col}13'].value,
                'number': ws[f'{col}12'].value
            }

        # The important data starts at this row
        start_row = 15
        max_row = 500  # Set a reasonable limit to avoid infinite loops

        # Define the columns to extract
        columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'Y']
        specific_sheet = ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']

        # Pre-load the data range we need to process
        data_range = []
        for row_idx in range(start_row, max_row + 1):
            row_values = {}
            for col in columns + specific_sheet:
                row_values[col] = ws[f'{col}{row_idx}'].value
            data_range.append(row_values)

            # Check if we've reached the end of data
            if row_values['B'] is None or row_values['B'] == '' or row_values['B'] == 'Shipment Method:':
                break

        # Process the pre-loaded data
        for row_data in data_range:
            if row_data['B'] is None or row_data['B'] == '' or row_data['B'] == 'Shipment Method:':
                break

            row = []
            for column in columns:
                cell_value = row_data[column]

                if column == 'G' and cell_value in matrix_codes:
                    matrix_id = matrix_codes[cell_value]
                    row.append(matrix_id)
                else:
                    row.append(cell_value)

            row.append(sampled_by)
            print(f"ANALISISSSSSSSSSSSSSSSSSSSS{analysis_requested}")
            row.append(analysis_requested)

            # Process specific sheets
            for sheet in specific_sheet:
                cell = row_data[sheet]

                if cell == 1:
                    sheet_name = sheet_headers[sheet]['name']
                    number_sheet = sheet_headers[sheet]['number']

                    final_sheet = f'{sheet_name} ({number_sheet})'
                    row.append(final_sheet)

            all_data.append(row)

        
        gc.collect()  # Force garbage collection

    except Exception as e:
        print(f"An error occurred in get_chain_data: {e}")
        return None

    return all_data


def get_matrix_data_flattened(wb, WSC,  route: str):
    """
    Process matrix data from Excel sheets and create flattened data structure.
    Each analysis gets its own complete row with sample identification data.
    """

    chain_data = get_chain_data(WSC)
    if chain_data is None or len(chain_data) == 0:
        return None

    HOJA_PREFIXES = [

        "Alkalinity", "Ammonia", "Apparent Color", "Chlorides",
        "Nitrates", "Nitrites", "Oil & Grease", "Ortho-phosphates",
        "Sulfate", "Total Dissolved Solids", "Nitrogen",
        "Total Hardness", "Phosphorous", "Total Solids",
        "Total Suspended Solids", "Turbidity"
    ]

    

    # Get all sheet names once
    sheets_in_excel = get_sheet_names(route)
    print(f"Hojas disponibles en el Excel: {sheets_in_excel}")

    # Create a dict for faster lookup of sample IDs
    sample_id_dict = {row[1]: idx for idx, row in enumerate(chain_data) if len(row) > 1}

    # Final flattened result list
    flattened_results = []

    try:
        # Dictionary to store sheet data to avoid reopening the same sheet
        sheet_cache = {}

        # Process each row in chain_data
        for i, row in enumerate(chain_data):
            if len(row) < 9:  # Ensure we have enough data
                continue

            sample_id = row[1]
            print(f"\nProcesando muestra: {sample_id}")

            # Extract the base sample data (common for all analyses)
            sample_id_data = row[:9]  # First 9 elements are identification data

            # --- PASO 1: Filtrar y identificar nombres de hojas del row original ---
            sheet_names_in_row = []

            for item in row:
                if (isinstance(item, str) and
                        any(item.startswith(prefix) for prefix in HOJA_PREFIXES) and
                        item in sheets_in_excel):
                    sheet_names_in_row.append(item)
                    print(f"  Hoja asociada encontrada: {item}")

            # If no relevant sheets found, keep the original row
            if not sheet_names_in_row:
                print(f"  No se encontraron hojas relevantes para la muestra {sample_id}")
                flattened_results.append(row)
                continue

            print(f"  Se procesarán {len(sheet_names_in_row)} hojas para la muestra {sample_id}")

            # --- PASO 2: Process each relevant sheet ---
            for sheet_name in sheet_names_in_row:
                print(f"  Procesando hoja: {sheet_name}")

                # Check if we already processed this sheet
                if sheet_name in sheet_cache:
                    sheet_data = sheet_cache[sheet_name]
                    print(f"    Usando datos en caché para {sheet_name}")
                else:
                    print(f"    Abriendo hoja {sheet_name}")
                    # Open the workbook with just the needed sheet

                    try:
                        ws = wb[sheet_name]

                        # Get constant values
                        analysis_requested = ws['M7'].value
                        cells_constant_value = ['N19', 'N20', 'N21', 'N22', 'N23', 'N24']
                        constantes = [ws[cell].value for cell in cells_constant_value]

                        # Pre-load all sample data from this sheet
                        samples_in_sheet = {}
                        start_row = 26
                        columns = ['B', 'C', 'D', 'E', 'F', 'H', 'I', 'J']

                        while True:
                            if ws[f'B{start_row}'].value == 'APPROVED BY':
                                break

                            current_sample_id = ws[f'C{start_row}'].value
                            if current_sample_id and current_sample_id in sample_id_dict:
                                # Extract all values except the sample ID (column C) which is redundant
                                sample_values = []
                                for column in columns:
                                    if column != 'C':  # Skip sample ID column
                                        sample_values.append(ws[f'{column}{start_row}'].value)

                                samples_in_sheet[current_sample_id] = sample_values
                                print(f"    Encontrada muestra {current_sample_id} en fila {start_row}")

                            start_row += 1
                            if start_row > 1000:  # Safety limit
                                break

                        sheet_data = {
                            'analysis_requested': analysis_requested,
                            'constantes': constantes,
                            'samples': samples_in_sheet
                        }

                        # Cache the sheet data
                        sheet_cache[sheet_name] = sheet_data

                    except Exception as e:
                        print(f"    Error procesando hoja {sheet_name}: {e}")
                        sheet_data = None

                    gc.collect()

                # Use the sheet data if available
                if sheet_data and sample_id in sheet_data['samples']:
                    # Create a new flattened row for this analysis
                    flattened_row = sample_id_data.copy()  # Start with sample identification data

                    # Add analysis data (removing the sample ID which is redundant)
                    analysis_data = [sheet_name, sheet_data['analysis_requested']] + sheet_data['constantes']

                    # Add sample specific measurements (excluding sample ID which is already in flattened_row)
                    sample_measurements = sheet_data['samples'][sample_id]

                    # Combine all data into one flattened row
                    flattened_row.extend(analysis_data + sample_measurements)

                    # Add this flattened row to results
                    flattened_results.append(flattened_row)
                    print(f"    Agregada fila aplanada con {len(flattened_row)} elementos para {sheet_name}")
                else:
                    print(f"    No se encontraron datos para la muestra {sample_id} en la hoja {sheet_name}")
        
    except Exception as e:
        print(f"Error general in get_matrix_data_flattened: {e}")
        return None

    return chain_data, flattened_results





