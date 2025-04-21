import os
from openpyxl import load_workbook

from scripts.excel.connect_excel import get_excel
from scripts.excel.get_sheet_names import get_sheet_names
from scripts.get.get_all_data import get_chain_data

# Prefijos para nombres de hoja
SHEET_PREFIXES = [
    "Alkalinity (471341)", "Ammonia (7664417)", "Apparent Color (471341)", "Chlorides (16887006)", "Nitrates (471341)",
    "Nitrites", "Oil & Grease (471341)", "Ortho-phosphates (471341)",
    "Sulfate (471341)", "Total Dissolved Solids (471341)", "Turbidity (471341)"
]

# Valores requeridos en columna B
REQUIRED_VALUES_B = {
    "Method Blank (MB)",
    "Laboratory Control Standard (LCS)",
    "QC",
    "Matrix Spike (MS)",
    "Matrix Spike Dup (MSD)"
}


def normalize_string(s):
    """Normaliza strings para comparación"""
    return " ".join(str(s).strip().split())


def generate_sample_id(base_id, increment):
    """Genera ID de muestra secuencial"""
    prefix, number = base_id.split('-')
    return f"{prefix}-{int(number) + increment:03d}"  # Formato 3 dígitos



def get_quality_data(wb, route: str):
    excel_sheets = get_sheet_names(route)
    cleaned_sheets = [sheet for sheet in excel_sheets if any(sheet.startswith(prefix) for prefix in SHEET_PREFIXES)]
    
    all_quality_controls = []
    columns = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
    position_constant_values = ["N19", "N20", "N21", "N22", "N23", "N24"]
    
    for sheet in cleaned_sheets:
        try:
            ws = wb[sheet]
            constant_values = [ws[cell].value for cell in position_constant_values]
            
            for row in range(1, 100):
                for col in ['B', 'C']:
                    cell_value = ws[f"{col}{row}"].value
                    if cell_value is None:
                        continue
                    
                    normalized_value = normalize_string(cell_value)
                    if any(qc_type in normalized_value for qc_type in REQUIRED_VALUES_B):
                        row_data = [sheet]
                        for col in columns:
                            cell_val = ws[f"{col}{row}"].value
                            # Filtrar fórmulas (opcional)
                            if isinstance(cell_val, str) and cell_val.startswith('='):
                                row_data.append(None)
                            else:
                                row_data.append(cell_val)
                        row_data.extend(constant_values)
                        all_quality_controls.append(row_data)
        except Exception as e:
            print(f"Error en hoja {sheet}: {e}")
    
    return all_quality_controls
 