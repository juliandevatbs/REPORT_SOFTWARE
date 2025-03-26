import os
import sys
from openpyxl import load_workbook

def get_analysis_requested():
    
    """
    Retrieves the value from cell M7 in the 'Chloride (16887006)' worksheet.
    
    Returns:
        The value from cell M7 or False if an error occurs.
    """
    sheet_name = "Chloride (16887006)"
    route = r"C:/Users/Duban Serrano/Desktop/REPORTES PYTHON/excel/Reporte 2025-03-12.xlsm"
    
    wb = None
    analysis_requested = False
    
    try:
        # Verify file exists
        if not os.path.exists(route):
            print("Error: File not found")
            return False
        
        # Open workbook
        wb = load_workbook(filename=route, read_only=True, data_only=True)
        
        # Verify sheet exists
        if sheet_name not in wb.sheetnames:
            print(f"Error: Worksheet '{sheet_name}' not found")
            return False
        
        # Get worksheet and cell value
        ws = wb[sheet_name]
        analysis_requested = ws['M7'].value
        
    except Exception as e:
        print(f"An error occurred: {e}")
        analysis_requested = False
        
    finally:
        # Clean up resources
        if wb is not None:
            try:
                wb.close()
            except Exception as e:
                print(f"Error closing workbook: {e}")
        
        return analysis_requested

get_analysis_requested()