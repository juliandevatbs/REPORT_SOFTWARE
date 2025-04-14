import os
import sys
from datetime import datetime
from openpyxl import load_workbook

from scripts.excel.connect_excel import get_excel

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from scripts.get.get_analysis_requested import get_analysis_requested

def get_lab_data(wb, ws):
    

    all_data = []
    
    try:
        
        # Extraer prefijo de AA3
        prefijo = str(ws['AA3'].value).strip() if ws['AA3'].value else ''
        
        # Define columns to extract (B, Y, C, D, E, G)
        columns = ['B', 'C', 'D', 'E', 'G', 'Y']
        start_row = 15
        
        # Contador para el consecutivo
        consecutivo = 1
        
        # Process rows until empty B column or 'Shipment Method:' is found
        for row in range(start_row, ws.max_row + 1):
            b_val = ws[f'B{row}'].value
            
            # Stop condition
            if b_val in (None, '', 'Shipment Method:'):
                break
            
            # Process valid row
            row_data = [len(all_data) + 1]  # Add sequential number
            
            # Extract values from specified columns
            for col in columns[1:]:  # Skip B column (already checked)
                cell = ws[f'{col}{row}']
                
                # Para la columna Y, generar consecutivo
                if col == 'Y':
                    # Generar número de serie con prefijo
                    numero_serie = f"{prefijo}-{consecutivo:03d}"
                    row_data.append(numero_serie)
                    consecutivo += 1
                else:
                    # Convert datetime if needed
                    if isinstance(cell.value, datetime):
                        row_data.append(cell.value)
                    else:
                        row_data.append(cell.value)
            
            # Add analysis requested (assuming this is constant per row)
            row_data.append(get_analysis_requested())
            
            all_data.append(row_data)
        
        print("All Data:", all_data)
        return all_data
    
    except Exception as e:
        print(f"Error: {str(e)}")
        # Print traceback for more detailed error information
        import traceback
        traceback.print_exc()
        return None


