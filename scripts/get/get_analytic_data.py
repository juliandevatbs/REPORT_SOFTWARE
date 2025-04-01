import os
from openpyxl import load_workbook

def get_analytic_data():
    """
    Extracts analytical data from the specified Excel worksheet.
    
    Returns:
        tuple: (all_data, constant_values) where:
            - all_data: List of lists containing row data from specified columns
            - constant_values: List of values from cells N21-N24
    """
    sheetname = 'Chloride (16887006)'
    route = r"C:/Users/Duban Serrano/Desktop/REPORTES PYTHON/excel/Reporte 2025-03-12 (3).xlsl"
    
    all_data = []
    constant_values = []
    wb = None
    
    try:
        # Verify file exists
        if not os.path.exists(route):
            print("Error: File not found")
            return [], []
        
        # Open workbook in read-only mode with data only (no formulas)
        wb = load_workbook(filename=route, read_only=True, data_only=True)
        
        # Verify sheet exists
        if sheetname not in wb.sheetnames:
            print(f"Error: Worksheet '{sheetname}' not found")
            return [], []
        
        ws = wb[sheetname]
        
        # The important data starts at this row
        row = 26
        
        # Columns with important data
        columns = ["C", "B", "F", "H"]
        
        # Get constant values
        constant_values = [
            ws['N24'].value,  # units_value
            ws['N21'].value,  # df_value
            ws['N22'].value,  # mdl_value
            ws['N23'].value   # pql_value
        ]
        
        while True:
            # Check if we should stop reading
            current_cell = ws[f'B{row}'].value
            if current_cell is None or current_cell == '' or current_cell == 'APPROVED BY':
                break
            
            # Extract data from all columns in current row
            data_row = []
            for column in columns:
                data_row.append(ws[f'{column}{row}'].value)
            
            all_data.append(data_row)
            row += 1
            
    except Exception as e:
        print(f"An error occurred: {e}")
        return [], []
        
    finally:
        # Clean up resources
        if wb is not None:
            try:
                wb.close()
            except Exception as e:
                print(f"Error closing workbook: {e}")
    
    return all_data, constant_values

def filter_none_data():
    """Legacy function kept for compatibility"""
    return get_analytic_data()

get_analytic_data()