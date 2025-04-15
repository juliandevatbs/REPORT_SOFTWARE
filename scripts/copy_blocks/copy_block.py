import traceback
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


def copy_range_with_styles(src_ws, dst_ws, src_range, dst_cell):
    try:
        print(f"Copiando rango {src_range} a celda {dst_cell}")

        # Obtener límites del rango fuente
        src_start_col, src_start_row, src_end_col, src_end_row = range_boundaries(src_range)

        # Obtener coordenadas destino
        dst_col_letter, dst_row = coordinate_from_string(dst_cell)
        dst_col = column_index_from_string(dst_col_letter)

        # Calcular dimensiones del rango a copiar
        width = src_end_col - src_start_col + 1
        height = src_end_row - src_start_row + 1

        print(f"Dimensiones: ancho={width}, alto={height}")

        # 1. Primero eliminar cualquier fusión existente en la zona de destino
        dst_end_col = dst_col + width - 1
        dst_end_row = dst_row + height - 1
        dst_range = f"{get_column_letter(dst_col)}{dst_row}:{get_column_letter(dst_end_col)}{dst_end_row}"

        # Guardar y eliminar fusiones existentes en el área de destino
        merges_to_remove = []
        for merged_range in dst_ws.merged_cells.ranges:
            m_range_str = str(merged_range)
            m_start_col, m_start_row, m_end_col, m_end_row = range_boundaries(m_range_str)

            # Verificar intersección con área de destino
            if not (m_end_row < dst_row or m_start_row > dst_end_row or
                    m_end_col < dst_col or m_start_col > dst_end_col):
                merges_to_remove.append(m_range_str)

        for merge_range in merges_to_remove:
            dst_ws.unmerge_cells(merge_range)

        print(f"Eliminadas {len(merges_to_remove)} fusiones existentes en el área de destino")

        # 2. Copiar celdas combinadas de la fuente
        merges_to_add = []
        for merged_range in src_ws.merged_cells.ranges:
            m_range_str = str(merged_range)
            m_start_col, m_start_row, m_end_col, m_end_row = range_boundaries(m_range_str)

            # Verificar si la celda combinada está dentro del rango a copiar
            if not (m_end_row < src_start_row or m_start_row > src_end_row or
                    m_end_col < src_start_col or m_start_col > src_end_col):
                # Calcular nueva posición en el destino
                new_start_col = m_start_col - src_start_col + dst_col
                new_start_row = m_start_row - src_start_row + dst_row
                new_end_col = m_end_col - src_start_col + dst_col
                new_end_row = m_end_row - src_start_row + dst_row

                new_merge = f"{get_column_letter(new_start_col)}{new_start_row}:{get_column_letter(new_end_col)}{new_end_row}"
                merges_to_add.append(new_merge)

        print(f"Preparadas {len(merges_to_add)} fusiones para agregar")

        # 3. Copiar dimensiones de columnas y filas primero para mantener tamaños
        # Copiar anchos de columnas
        for col_idx in range(src_start_col, src_end_col + 1):
            src_col = get_column_letter(col_idx)
            dst_col_idx = col_idx - src_start_col + dst_col
            dst_col_letter = get_column_letter(dst_col_idx)

            if src_col in src_ws.column_dimensions:
                src_dim = src_ws.column_dimensions[src_col]
                dst_ws.column_dimensions[dst_col_letter].width = src_dim.width
                dst_ws.column_dimensions[dst_col_letter].hidden = src_dim.hidden
                dst_ws.column_dimensions[dst_col_letter].outline_level = src_dim.outline_level

        # Copiar alturas de filas
        for row_idx in range(src_start_row, src_end_row + 1):
            dst_row_idx = row_idx - src_start_row + dst_row

            if row_idx in src_ws.row_dimensions:
                src_dim = src_ws.row_dimensions[row_idx]
                dst_ws.row_dimensions[dst_row_idx].height = src_dim.height
                dst_ws.row_dimensions[dst_row_idx].hidden = src_dim.hidden
                dst_ws.row_dimensions[dst_row_idx].outline_level = src_dim.outline_level

        print("Dimensiones de filas y columnas copiadas")

        # 4. Copiar valores y estilos celda por celda
        for row_idx in range(src_start_row, src_end_row + 1):
            for col_idx in range(src_start_col, src_end_col + 1):
                src_cell = src_ws.cell(row=row_idx, column=col_idx)

                # Calcular coordenadas destino
                dst_row_idx = row_idx - src_start_row + dst_row
                dst_col_idx = col_idx - src_start_col + dst_col
                dst_cell = dst_ws.cell(row=dst_row_idx, column=dst_col_idx)

                # Copiar valor
                dst_cell.value = src_cell.value

                # Copiar estilos
                if src_cell.has_style:
                    # Fuente
                    if src_cell.font:
                        dst_cell.font = Font(
                            name=src_cell.font.name,
                            size=src_cell.font.size,
                            bold=src_cell.font.bold,
                            italic=src_cell.font.italic,
                            vertAlign=src_cell.font.vertAlign,
                            underline=src_cell.font.underline,
                            strike=src_cell.font.strike,
                            color=src_cell.font.color
                        )

                    # Relleno
                    if src_cell.fill and hasattr(src_cell.fill, 'fill_type') and src_cell.fill.fill_type:
                        dst_cell.fill = PatternFill(
                            fill_type=src_cell.fill.fill_type,
                            start_color=src_cell.fill.start_color,
                            end_color=src_cell.fill.end_color
                        )

                    # Bordes
                    if src_cell.border:
                        dst_cell.border = Border(
                            left=Side(style=src_cell.border.left.style,
                                      color=src_cell.border.left.color) if src_cell.border.left and src_cell.border.left.style else None,
                            right=Side(style=src_cell.border.right.style,
                                       color=src_cell.border.right.color) if src_cell.border.right and src_cell.border.right.style else None,
                            top=Side(style=src_cell.border.top.style,
                                     color=src_cell.border.top.color) if src_cell.border.top and src_cell.border.top.style else None,
                            bottom=Side(style=src_cell.border.bottom.style,
                                        color=src_cell.border.bottom.color) if src_cell.border.bottom and src_cell.border.bottom.style else None
                        )

                    # Alineación
                    if src_cell.alignment:
                        dst_cell.alignment = Alignment(
                            horizontal=src_cell.alignment.horizontal,
                            vertical=src_cell.alignment.vertical,
                            textRotation=src_cell.alignment.textRotation,
                            wrapText=src_cell.alignment.wrapText,
                            shrinkToFit=src_cell.alignment.shrinkToFit,
                            indent=src_cell.alignment.indent
                        )

                    # Formato numérico
                    dst_cell.number_format = src_cell.number_format

        print("Valores y estilos copiados correctamente")

        # 5. Aplicar las fusiones de celdas después de copiar valores y estilos
        for merge_range in merges_to_add:
            dst_ws.merge_cells(merge_range)

        print(f"Aplicadas {len(merges_to_add)} fusiones de celdas")

        return True

    except Exception as e:
        print(f"Error durante la copia: {str(e)}")
        traceback.print_exc()
        return False