import datetime
import traceback
import customtkinter
import os
import threading
from PIL import Image
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from datetime import datetime
# Import the execute_all function - modify the import path as needed


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.title("Report Generator")
        self.geometry("900x600")
        self.current_file = None
        self.sheet_buttons = []
        self.current_sheet = None

        # Configurar grid principal
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # Cargar imágenes
        image_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "assets\icons")
        self.logo_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "LOGO SRL FINAL.png")),
                                                 size=(30, 30))
        self.large_test_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "LOGO SRL FINAL.png")),
                                                       size=(150, 150))
        self.image_icon_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "file-type-xls.png")),
                                                       size=(20, 20))
        self.sheet_icon = customtkinter.CTkImage(Image.open(os.path.join(image_path, "file-type-xls.png")),
                                                 size=(15, 15))

        # Frame de navegación izquierdo
        self.navigation_frame = customtkinter.CTkFrame(self, corner_radius=0, width=150)
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")
        self.navigation_frame.grid_rowconfigure(5, weight=1)

        self.navigation_frame_label = customtkinter.CTkLabel(self.navigation_frame, text="  Report Generator",
                                                             image=self.logo_image,
                                                             compound="left",
                                                             font=customtkinter.CTkFont(size=15, weight="bold"))
        self.navigation_frame_label.grid(row=0, column=0, padx=10, pady=20)

        self.home_button = customtkinter.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10,
                                                   text="Home",
                                                   fg_color="transparent", text_color=("gray10", "gray90"),
                                                   hover_color=("gray70", "gray30"),
                                                   image=self.image_icon_image, anchor="w",
                                                   command=self.home_button_event)
        self.home_button.grid(row=1, column=0, sticky="ew", padx=5)

        self.history_button = customtkinter.CTkButton(self.navigation_frame, corner_radius=0, height=40,
                                                      border_spacing=10, text="History",
                                                      fg_color="transparent", text_color=("gray10", "gray90"),
                                                      hover_color=("gray70", "gray30"),
                                                      image=self.image_icon_image, anchor="w")
        self.history_button.grid(row=2, column=0, sticky="ew", padx=5)

        self.settings_button = customtkinter.CTkButton(self.navigation_frame, corner_radius=0, height=40,
                                                       border_spacing=10, text="Settings",
                                                       fg_color="transparent", text_color=("gray10", "gray90"),
                                                       hover_color=("gray70", "gray30"),
                                                       image=self.image_icon_image, anchor="w")
        self.settings_button.grid(row=3, column=0, sticky="ew", padx=5)

        self.separator = customtkinter.CTkFrame(self.navigation_frame, height=2, fg_color="gray50")
        self.separator.grid(row=4, column=0, padx=10, pady=10, sticky="ew")

        # Menú de apariencia
        self.appearance_mode_menu = customtkinter.CTkOptionMenu(self.navigation_frame,
                                                                values=["Light", "Dark", "System"],
                                                                command=self.change_appearance_mode_event)
        self.appearance_mode_menu.grid(row=6, column=0, padx=10, pady=20, sticky="s")

        # Frame principal (home)
        self.home_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.home_frame.grid_columnconfigure(0, weight=1)
        self.home_frame.grid_rowconfigure(1, weight=1)

        # Logo en la parte superior
        self.home_frame_large_image_label = customtkinter.CTkLabel(self.home_frame, text="",
                                                                   image=self.large_test_image)
        self.home_frame_large_image_label.grid(row=0, column=0, padx=20, pady=10)

        # Contenedor para el selector de archivo y el grid de hojas
        self.main_container = customtkinter.CTkFrame(self.home_frame, fg_color="transparent")
        self.main_container.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        self.main_container.grid_columnconfigure(0, weight=1)
        self.main_container.grid_rowconfigure(1, weight=1)

        # Componentes para selección de archivo
        self.file_selection_frame = customtkinter.CTkFrame(self.main_container, fg_color="transparent")
        self.file_selection_frame.grid(row=0, column=0, sticky="ew", pady=(0, 20))

        self.selected_file_label = customtkinter.CTkLabel(self.file_selection_frame, text="No file selected",
                                                          font=customtkinter.CTkFont(size=12))
        self.selected_file_label.pack(side="left", padx=(0, 10))

        # Contenedor para el botón y la X
        self.select_file_container = customtkinter.CTkFrame(self.file_selection_frame, fg_color="transparent")
        self.select_file_container.pack(side="right")

        # Botón principal
        self.select_file_button = customtkinter.CTkButton(
            self.select_file_container,
            text="Choose Excel file",
            image=self.image_icon_image,
            command=self.select_excel_file
        )
        self.select_file_button.pack(side="left")

        # Botón X (inicialmente oculto)
        self.close_button = customtkinter.CTkButton(
            self.select_file_container,
            text="✕",
            width=28,
            height=28,
            fg_color="transparent",
            hover_color=("gray70", "gray30"),
            command=self.clear_file_selection
        )
        self.close_button.pack(side="left", padx=(5, 0))
        self.close_button.pack_forget()

        self.sheets_grid_frame = customtkinter.CTkScrollableFrame(
            self.main_container,
            fg_color="transparent"
        )

        self.sheets_grid_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        self.sheets_grid_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)

        # Label inicial para cuando no hay archivo seleccionado
        self.no_sheets_label = customtkinter.CTkLabel(
            self.sheets_grid_frame,
            text="Select an Excel file to view its sheets",
            font=customtkinter.CTkFont(size=12, slant="italic")
        )
        self.no_sheets_label.grid(row=1, column=0, pady=50)

        # Crear un frame para el loading indicator
        self.loading_frame = customtkinter.CTkFrame(self.home_frame, fg_color="transparent")
        self.loading_frame.grid(row=2, column=0, sticky="ew", pady=10)

        # Progress bar para mostrar que se está trabajando en el reporte
        self.progress_bar = customtkinter.CTkProgressBar(self.loading_frame)
        self.progress_bar.grid(row=0, column=0, padx=20, pady=10, sticky="ew")
        self.progress_bar.set(0)
        self.progress_bar.grid_remove()  # Ocultar inicialmente

        # Label para mostrar estado
        self.status_label = customtkinter.CTkLabel(self.loading_frame, text="")
        self.status_label.grid(row=1, column=0, padx=20, pady=5)
        self.status_label.grid_remove()  # Ocultar inicialmente

        # Seleccionar frame inicial
        self.select_frame_by_name("home")

    def _update_progress(self, message, progress=None):
        """Update progress bar and status label"""
        if progress is not None:
            # Si la barra está en modo indeterminado, cámbiala a determinado
            if self.progress_bar.cget("mode") == "indeterminate":
                self.progress_bar.stop()
                self.progress_bar.configure(mode="determinate")
            self.progress_bar.set(progress)

        if message:
            self.status_label.configure(text=message)

    def _do_update_progress(self, value, status_text):
        if value is not None:
            # Cambiar a modo determinado si es necesario
            if self.progress_bar.cget("mode") == "indeterminate":
                self.progress_bar.stop()
                self.progress_bar.configure(mode="determinate")
            self.progress_bar.set(value)

        if status_text:
            self.status_label.configure(text=status_text)

    def display_sheets(self, file_path):
        """Display Excel sheets in a grid layout"""
        # Limpiar el frame
        for widget in self.sheets_grid_frame.winfo_children():
            widget.destroy()

        try:
            # Leer las hojas del archivo Excel
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheets = workbook.sheetnames

            if not sheets:
                self.no_sheets_label = customtkinter.CTkLabel(
                    self.sheets_grid_frame,
                    text="The selected Excel file has no sheets",
                    font=customtkinter.CTkFont(size=12)
                )
                self.no_sheets_label.grid(row=0, column=0, pady=50)
                return

            # Configurar el grid
            self.sheets_grid_frame.grid_columnconfigure(0, weight=1)
            self.sheets_grid_frame.grid_columnconfigure(1, weight=1)
            self.sheets_grid_frame.grid_columnconfigure(2, weight=1)

            # Crear un grid de 3 columnas
            cols = 3
            for i, sheet in enumerate(sheets):
                row = i // cols
                col = i % cols

                sheet_btn = customtkinter.CTkButton(
                    self.sheets_grid_frame,
                    text=sheet,
                    image=self.sheet_icon,
                    compound="top",
                    width=150,
                    height=150,
                    corner_radius=10,
                    fg_color=("gray85", "gray25"),
                    hover_color=("gray75", "gray35"),
                    command=lambda s=sheet: self.sheet_selected(s)
                )
                sheet_btn.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")

        except Exception as e:
            error_msg = f"Could not read sheets:\n{str(e)}"
            print(error_msg)  # Para depuración
            messagebox.showerror("Error", error_msg)
            self.no_sheets_label = customtkinter.CTkLabel(
                self.sheets_grid_frame,
                text="Error loading sheets",
                font=customtkinter.CTkFont(size=12)
            )
            self.no_sheets_label.grid(row=0, column=0, pady=50)

    def select_excel_file(self):
        """Open file dialog to select Excel file"""
        filetypes = (
            ('Excel files', '*.xlsx *.xls'),
            ('All files', '*.*')
        )

        file_path = filedialog.askopenfilename(
            title='Select Excel file',
            initialdir='/',
            filetypes=filetypes
        )

        if file_path:
            self.current_file = file_path
            file_name = os.path.basename(file_path)
            self.selected_file_label.configure(text=f"Selected: {file_name}")

            confirm = messagebox.askyesno(
                "Confirm Upload",
                f"Do you want to upload this file?\n\n{file_path}",
                parent=self
            )

            if confirm:
                try:
                    self.upload_excel_file(file_path)
                    self.display_sheets(file_path)
                    # Actualizar botón SOLO después de mostrar las hojas
                    self.select_file_button.configure(text="Generar Reporte",
                                                      command=lambda: self._execute_report_thread())
                    self.close_button.pack(side="left", padx=(5, 0))
                except Exception as e:
                    messagebox.showerror("Error", f"Error processing file:\n{str(e)}")
                    self.clear_file_selection()

    def clear_file_selection(self):
        """Clear the current file selection"""
        self.current_file = None
        self.current_sheet = None
        self.selected_file_label.configure(text="No file selected")
        self.select_file_button.configure(text="Choose Excel file")
        self.close_button.pack_forget()

        # Limpiar el grid de hojas
        for widget in self.sheets_grid_frame.winfo_children():
            widget.destroy()

        # Mostrar mensaje inicial
        self.no_sheets_label = customtkinter.CTkLabel(
            self.sheets_grid_frame,
            text="Select an Excel file to view its sheets",
            font=customtkinter.CTkFont(size=12, slant="italic")
        )
        self.no_sheets_label.grid(row=0, column=0, pady=50)

    def display_sheets(self, file_path):
        """Display Excel sheets in a grid layout"""
        # Limpiar el frame
        for widget in self.sheets_grid_frame.winfo_children():
            widget.destroy()

        try:
            # Leer las hojas del archivo Excel
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheets = workbook.sheetnames

            if not sheets:
                self.no_sheets_label = customtkinter.CTkLabel(
                    self.sheets_grid_frame,
                    text="The selected Excel file has no sheets",
                    font=customtkinter.CTkFont(size=12)
                )
                self.no_sheets_label.grid(row=0, column=0, pady=50)
                return

            # Configurar el grid
            self.sheets_grid_frame.grid_columnconfigure(0, weight=1)
            self.sheets_grid_frame.grid_columnconfigure(1, weight=1)
            self.sheets_grid_frame.grid_columnconfigure(2, weight=1)

            # Crear un grid de 3 columnas
            cols = 3
            for i, sheet in enumerate(sheets):
                row = i // cols
                col = i % cols

                sheet_btn = customtkinter.CTkButton(
                    self.sheets_grid_frame,
                    text=sheet,
                    image=self.sheet_icon,
                    compound="top",
                    width=150,
                    height=150,
                    corner_radius=10,
                    fg_color=("gray85", "gray25"),
                    hover_color=("gray75", "gray35"),
                    command=lambda s=sheet: self.sheet_selected(s)
                )
                sheet_btn.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")

        except Exception as e:
            error_msg = f"Could not read sheets:\n{str(e)}"
            print(error_msg)  # Para depuración
            messagebox.showerror("Error", error_msg)
            self.no_sheets_label = customtkinter.CTkLabel(
                self.sheets_grid_frame,
                text="Error loading sheets",
                font=customtkinter.CTkFont(size=12)
            )
            self.no_sheets_label.grid(row=0, column=0, pady=50)

    def sheet_selected(self, sheet_name):
        """Handle sheet selection"""
        self.current_sheet = sheet_name
        messagebox.showinfo("Sheet Selected",
                            f"Sheet selected: {sheet_name}\n\nFile: {os.path.basename(self.current_file)}")
        # Aquí puedes agregar la lógica para generar el reporte

    def upload_excel_file(self, file_path):
        """Process the uploaded Excel file"""
        try:
            # Aquí iría tu lógica real de procesamiento del archivo
            messagebox.showinfo(
                "Success",
                f"File uploaded successfully!\n\n{file_path}",
                parent=self
            )

            self.selected_file_label.configure(text=f"Uploaded: {os.path.basename(file_path)}")

        except Exception as e:
            messagebox.showerror(
                "Error",
                f"Failed to upload file:\n\n{str(e)}",
                parent=self
            )
            self.selected_file_label.configure(text="Error uploading file")
            self.clear_file_selection()

    def select_frame_by_name(self, name):
        # Cambiar color del botón activo
        self.home_button.configure(fg_color=("gray75", "gray25") if name == "home" else "transparent")
        self.history_button.configure(fg_color=("gray75", "gray25") if name == "history" else "transparent")
        self.settings_button.configure(fg_color=("gray75", "gray25") if name == "settings" else "transparent")

        # Mostrar el frame correspondiente
        if name == "home":
            self.home_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.home_frame.grid_forget()

    def home_button_event(self):
        self.select_frame_by_name("home")

    def change_appearance_mode_event(self, new_appearance_mode):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def _execute_report_thread(self):
        """Thread worker function to generate report"""
        try:
            # Actualizar la UI periódicamente
            def update_progress(status):
                self.after(0, lambda: self.status_label.configure(text=status))

            # Configurar puntos de actualización en el proceso
            update_progress("Procesando datos...")

            # Usar una cola para comunicación entre hilos
            from queue import Queue
            status_queue = Queue()

            # Modificar execute_all para que reciba la cola y envíe actualizaciones
            from execute_all import execute_all
            success = execute_all(status_queue=status_queue)

            # Revisar periódicamente las actualizaciones de estado
            while not status_queue.empty():
                status = status_queue.get()
                update_progress(status)



        except Exception as e:
            print("N")
    # In the App class, modify the execute_report_thread method to pass the file path:

    def _execute_report_thread(self):
        """Thread worker function to generate report"""
        # Mostrar la barra de progreso y configurarla
        self.progress_bar.grid()
        self.status_label.grid()
        self.progress_bar.configure(mode="indeterminate")
        self.progress_bar.start()
        self.status_label.configure(text="Generando reporte...", text_color="white")

        # Deshabilitar botones durante la generación
        self.select_file_button.configure(state="disabled")
        self.close_button.configure(state="disabled")

        # Ejecutar en un hilo separado
        threading.Thread(target=self._generate_report_background, daemon=True).start()

    def _generate_report_background(self):
        """Background task to generate the report"""
        try:
            # Definir callback para actualizaciones
            def update_progress(message, progress=None):
                self.after(0, lambda: self._update_progress_internal(message, progress))

            # Importar y ejecutar la función de generación de reportes
            from execute_all import execute_all
            success = execute_all(status_callback=update_progress)

            # Mostrar resultado
            if success:
                self.after(0, self._report_complete, True, "Reporte generado exitosamente!")
            else:
                self.after(0, self._report_complete, False, "Error al generar el reporte")

        except Exception as e:
            error_message = f"Error al generar reporte:\n\n{str(e)}"
            traceback_str = traceback.format_exc()
            print(traceback_str)
            self.after(0, self._report_complete, False, error_message)

    def _update_progress_internal(self, message, progress=None):
        """Internal method to update progress (called from main thread)"""
        if progress is not None:
            if self.progress_bar.cget("mode") == "indeterminate":
                self.progress_bar.stop()
                self.progress_bar.configure(mode="determinate")
            self.progress_bar.set(progress)

        if message:
            self.status_label.configure(text=message)

    def _report_complete(self, success, message):
        """Update UI after report generation is complete"""
        # Detener y ocultar la barra de progreso
        self.progress_bar.stop()
        self.progress_bar.grid_remove()
        self.status_label.grid_remove()

        # Restaurar estado de los botones
        self.select_file_button.configure(state="normal")
        self.close_button.configure(state="normal")

        # Mostrar mensaje
        if success:
            messagebox.showinfo("Success", message, parent=self)
        else:
            messagebox.showerror("Error", message, parent=self)



if __name__ == "__main__":
    app = App()
    app.mainloop()